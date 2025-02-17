from elasticsearch import Elasticsearch
from elasticsearch.exceptions import RequestError, ConnectionError, TransportError, ApiError
from datetime import datetime
import os
import pandas as pd
import time
from openpyxl import load_workbook, Workbook
from pathlib import Path

ELASTIC_HOST = "http://172.10.12.31:9200"
ELASTIC_USER = "root"
ELASTIC_PASSWORD = "uzZS6e2rhf56"
BASE_DIR = "D:\\code\\dir"
sa_file_anme = "Sa_host_message_file"


# Создаем директорию один раз
os.makedirs(BASE_DIR, exist_ok=True)

# Инициализируем клиент Elasticsearch
es = Elasticsearch(ELASTIC_HOST, basic_auth=(ELASTIC_USER, ELASTIC_PASSWORD))


def load_processed_indices(filename='indexes.txt'):
    processed_indices = set()
    if os.path.exists(filename):
        try:
            with open(filename, 'r', encoding='utf-8') as f:
                processed_indices = {line.strip() for line in f if line.strip()}
        except IOError as e:
            print(f"Ошибка чтения файла {filename}: {str(e)}")
    return processed_indices

def save_processed_index(filename='indexes.txt', index_name=None):
    try:
        with open(filename, 'a', encoding='utf-8') as f:
            f.write(f"{index_name}\n")
    except IOError as e:
        print(f"Ошибка записи в файл {filename}: {str(e)}")

def append_to_excel_optimized(data_list):
    full_path = os.path.join(BASE_DIR, sa_file_anme)
    
    try:
        # Проверяем существование файла и создаём новый если нужно
        if not os.path.exists(full_path):
            wb = Workbook()
            ws = wb.active
            ws.title = sa_file_anme
            
            headers = ['Хост', 'Пользователь', 'Индекс', 
                      'Короткое сообщение', 'Сообщение']
            ws.append(headers)
            
            wb.save(full_path)
        
        # Загружаем рабочую книгу
        wb = load_workbook(full_path)
        ws = wb[sa_file_anme]
        
        # Преобразуем данные в DataFrame для эффективной обработки
        df_new = pd.DataFrame(data_list)
        
        # Добавляем данные в лист
        for _, row in df_new.iterrows():
            ws.append([
                row['Хост'],
                row['Пользователь'],
                row['Индекс'],
                row['Короткое сообщение'],
                row['Сообщение']
            ])
        
        # Сохраняем изменения
        wb.save(full_path)
        print(f"Данные успешно добавлены в файл {full_path}")
        return True
        
    except Exception as e:
        print(f"Ошибка при добавлении данных: {str(e)}")
        return False
    finally:
        if 'wb' in locals():
            wb.close()

def process_indices(es, host_name):
    try:
        # Получаем список всех индексов
        indices_response = es.cat.indices(h='index', s='index')
        processed_indices = load_processed_indices()
        new_indices = sorted(set(indices_response.split()) - processed_indices)
        
        if not new_indices:
            print("[{}] Нет новых индексов для обработки".format(datetime.now()))
            return
            
        batch_size = 10000
        buffer_data = []
        
        
        for index_name in new_indices:
            try:
                
                body = {
                    "size": batch_size,
                    "_source": ["message", "host.name", "winlog.event_data.SubjectUserName"],
                    "query": {"bool": {"must": [{"match": {"winlog.event_data.SubjectUserName": "sa"}}]}}
                }
                
                buffer_data = pd.DataFrame(columns=[
                        'Хост',
                        'Пользователь',
                        'Индекс',
                        'Короткое сообщение',
                        'Сообщение'
                    ])
                index_name = "winpc-2024.12.12"
                # Используем bulk API для эффективной обработки
                response = es.search(index=index_name, body=body, scroll='5m')
                scroll_id = response['_scroll_id']
                print(f"[{datetime.now()}] Начало обработки документов в индексе {index_name}")
                
                while True:
                    hits = response['hits']['hits']
                    if not hits:
                        break
                        
                    for hit in hits:
                        try:
                            # Извлечение данных с безопасным доступом
                            host_name = hit.get('_source', {}).get('host', {}).get('name')
                            message = hit.get('_source', {}).get('message')
                            username = hit.get('_source', {}).get('winlog', {}).get('event_data', {}).get('SubjectUserName')
                            
                            # Проверка наличия обязательных полей
                            if not all([host_name, message]):
                                print(f"[{datetime.now()}] Предупреждение: отсутствуют обязательные поля в документе")
                                continue
                                
                            # Обработка сообщения
                            first_sentence = message.split('.')[0].strip() if '.' in message else message
                            
                            # Создание DataFrame
                            df_buffer = pd.DataFrame({
                                'Хост': [host_name],
                                'Пользователь': [username],
                                'Индекс': [index_name],
                                'Короткое сообщение': [first_sentence],
                                'Сообщение': [message]
                            })
                            
                            # Безопасное объединение DataFrame
                            if buffer_data is not None:
                                if not buffer_data.empty:  # Проверяем, что DataFrame не пустой
                                    buffer_data = pd.concat([buffer_data, df_buffer], ignore_index=True)
                                else:
                                    buffer_data = df_buffer
                            else:
                                buffer_data = df_buffer
                                
                        except KeyError as e:
                            print(f"[{datetime.now()}] Ошибка доступа к ключу: {str(e)}")
                            continue
                        except ValueError as e:
                            print(f"[{datetime.now()}] Ошибка обработки значения: {str(e)}")
                            continue
                        except Exception as e:
                            print(f"[{datetime.now()}] Неожиданная ошибка: {str(e)}")
                            continue
                                        
                    # Проверка завершения скроллинга
                    if len(response['hits']['hits']) == 0:
                        break
                        
                    response = es.scroll(scroll_id=scroll_id, scroll='5m')
                
                # Сохраняем индекс как обработанный
                save_processed_index(index_name=index_name)
                print(f"[{datetime.now()}] Обработано : {index_name}")
                
                # Записываем данные в Excel
                if buffer_data:
                    full_df = pd.concat(buffer_data, ignore_index=True)
                    
                    success = append_to_excel_optimized(full_df.values.tolist())
                    if success:
                        print(f"Данные успешно записаны ")
                    else:
                        print(f"Ошибка при записи данных")
                
                # Очищаем буфер
                buffer_data.clear()
                
            except TransportError as e:
                print(f"Ошибка транспорта при обработке индекса {index_name}: {str(e)}")
            except RequestError as e:
                print(f"Ошибка запроса при обработке индекса {index_name}: {str(e)}")
            except ConnectionError as e:
                print(f"Ошибка подключения при обработке индекса {index_name}: {str(e)}")
            except Exception as e:
                print(f"Ошибка при обработке индекса {index_name}: {str(e)}")
                
    except Exception as e:
        print(f"Критическая ошибка при обработке индексов: {str(e)}")
        

process_indices(es, sa_file_anme)


Ошибка при обработке индекса winpc-2024.12.12: The truth value of a DataFrame is ambiguous. Use a.empty, a.bool(), a.item(), a.any() or a.all().
