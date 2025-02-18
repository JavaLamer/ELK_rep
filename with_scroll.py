from elasticsearch import Elasticsearch
from elasticsearch.exceptions import RequestError, ConnectionError, TransportError, ApiError
from datetime import datetime
import os
import pandas as pd
import time
from openpyxl import load_workbook, Workbook
from pathlib import Path



ELASTIC_HOST = "http://172.10.12.31:9200"
ELASTIC_USER = "root"
ELASTIC_PASSWORD = "uzZS6e2rhf56"

BASE_DIR = "D:\\code\\dir"
os.makedirs(BASE_DIR, exist_ok=True)

es = Elasticsearch(ELASTIC_HOST, basic_auth=(ELASTIC_USER, ELASTIC_PASSWORD))
size=5000

def get_hostname():
    host_name = "ERIPAVC2022.rrb.by"

    return host_name


def create_excel_template(filename, index):
    full_path = os.path.join(BASE_DIR, filename)
    
    if os.path.exists(full_path):
        return
    
    try:
        with pd.ExcelWriter(full_path, mode='w', engine='openpyxl') as writer:
            df = pd.DataFrame(columns=[
                'Хост',
                'Пользователь',
                'Индекс',
                'Короткое сообщение',
                'Сообщение'
            ])
            df.to_excel(writer, sheet_name=get_hostname(), index=False)
    except Exception as e:
        print(f"Ошибка при создании файла: {e}")
      
        
def normalize_text(text):
    return text.lower().strip()


def load_processed_indices(filename='indexes.txt'):
    processed_indices = set()
    if os.path.exists(filename):
        try:
            with open(filename, 'r', encoding='utf-8') as f:
                processed_indices = {line.strip() for line in f if line.strip()}
        except IOError as e:
            print(f"Ошибка чтения файла {filename}: {str(e)}")
    return processed_indices

def save_processed_index(filename='indexes.txt', index_name=None):
    try:
        with open(filename, 'a', encoding='utf-8') as f:
            f.write(f"{index_name}\n")
    except IOError as e:
        print(f"Ошибка записи в файл {filename}: {str(e)}")

def append_to_excel(data_list, filename):
    full_path = os.path.join(BASE_DIR, filename)
    try:
        if not os.path.exists(full_path):
            raise FileNotFoundError(f"Файл {full_path} не существует. Сначала создайте шаблон.")
        
        # Загружаем существующий файл
        existing_wb = load_workbook(full_path)
        sheet_name = get_hostname()
        ws_source = existing_wb[sheet_name]
        
        # Добавляем новые данные напрямую в существующий файл
        df = pd.DataFrame(data_list)
        start_row = ws_source.max_row + 1
        for idx, row in df.iterrows():
            ws_source.append(list(row))
        
        # Сохраняем изменения
        existing_wb.save(full_path)
        print(f"Данные успешно добавлены в файл {full_path}")
        return True
        
    except Exception as e:
        print(f"Ошибка при добавлении данных: {str(e)}")
        return False
    finally:
        # Закрываем рабочую книгу
        if 'existing_wb' in locals():
            existing_wb.close()


def process_indices(es, host_name):
    try:
        if not es.ping():
            raise ConnectionError("Нет соединения с Elasticsearch")
            
        body = {
    "size": size,
    "_source": ["message", "host.name", "winlog.event_data.SubjectUserName"],
    "query": {
        "bool": {
            "must": [
                { "match": { "winlog.event_data.SubjectUserName": "sa" }}
                
      ]
    }
  }
}
        
        processed_indices = load_processed_indices()
        unique_combinations = {}
        excel_data_unique = []
        excel_data_full = []
        base_filename = f"{get_hostname()}_excel"
        
        indices_response = es.cat.indices(h='index', s='index')
        
        unique_filename = f"{base_filename}_unique.xlsx"
       
        
        
        for index_name in indices_response.split():
            if index_name in processed_indices:
                print(f"Пропуск индекса {index_name} (уже обработан)")
                continue
                
            create_excel_template(unique_filename, index= index_name)
               
            
            try:
                response = es.search(
                    index=index_name,
                    body=body,
                    scroll='1m'
                )
                

                while True:
                    for hit in response['hits']['hits']:
                        try:
                            host_name = hit['_source']['host']['name']
                            message = hit['_source']['message']
                            username = hit.get('_source', {}).get('winlog', {}).get('event_data', {}).get('SubjectUserName')
                            
                    
                            
                            first_sentence = message.split('.')[0].strip() if '.' in message else message
                            combination_key = normalize_text(first_sentence)
                            

                            if combination_key not in unique_combinations:
                                print(f"{first_sentence}\n{combination_key}\n\n")
                                unique_combinations[combination_key] = True
                                excel_data_unique.append({
                                        'Хост': host_name,
                                        'Пользователь': username,
                                        'Индекс': index_name,
                                        'Короткое сообщение': first_sentence,
                                        'Сообщение': message
                                        
                                        
                                })
                                append_to_excel(excel_data_unique, unique_filename)
                                print(f"* - {index_name}")

                            
                            
                        except KeyError as e:
                            print(f"Ошибка при обработке записи winlog в индексе {index_name}: {str(e)}")
                            break  
                            
                    if len(response['hits']['hits']) == 0:
                        print(f"Пустой - {index_name}")
                        break
                        
                response = es.scroll(scroll_id=response['_scroll_id'], scroll='1m')
                save_processed_index(index_name=index_name)
                continue

                    
            except ApiError as e:
                print(f"Ошибка при обработке индекса {index_name}: {str(e)}")
                time.sleep(50)
                pass
                
    except ConnectionError as e:
        print(f"Ошибка подключения к Elasticsearch: {str(e)}")
    except Exception as e:
        print(f"Неожиданная ошибка: {str(e)}")

# Запускаем обработку
host_name = get_hostname()
process_indices(es, host_name)

